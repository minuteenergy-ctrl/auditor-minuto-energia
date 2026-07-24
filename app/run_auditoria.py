# -*- coding: utf-8 -*-
"""
run_auditoria.py -- Runner unificado de auditoria de faturas
Uso:
    python run_auditoria.py --concessionaria neoenergia_pe  --pasta ./faturas/
    python run_auditoria.py --concessionaria cpfl_piratininga --pasta ./faturas/
    python run_auditoria.py --concessionaria todas --pasta_neo ./fat_neo/ --pasta_cpfl ./fat_cpfl/

Saida: output/run_YYYYMMDD_HHMMSS/
  auditoria_mestre.xlsx      (RESUMO / MESTRE / OK / INVESTIGAR / DIVERGENCIA)
  individuais/*.xlsx
"""
import argparse
import datetime
import os
import sys
from pathlib import Path

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

# -- CONFIG padrao CPFL -------------------------------------------------------
CPFL_CONFIG_DEFAULT = {
    "data_adesao_mmgd": "2022-01-01",
    "usar_cat_como_compensada": True,
}


# -- Helpers ------------------------------------------------------------------

def _fmt_date(v):
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    return str(v)


# -- Normalizadores -----------------------------------------------------------

def normalizar_neoenergia_pe(rec, triagem, motivos, metricas):
    return {
        "arquivo":        rec.get("arquivo"),
        "distribuidora":  "Neoenergia PE",
        "layout":         rec.get("layout"),
        "ref_mes_ano":    rec.get("ref_mes_ano"),
        "vencimento":     rec.get("vencimento"),
        "conta_uc":       rec.get("conta_contrato"),
        "cliente_nome":   None,
        "subgrupo":       "B3",
        "data_emissao":   rec.get("data_emissao"),
        "nr_nota_fiscal": rec.get("nr_nota_fiscal"),
        "nr_medidor":     rec.get("nr_medidor"),
        "consumo_kwh":    rec.get("consumo_kwh_tusd_qtd"),
        "nr_dias":        rec.get("nr_dias"),
        "leitura_anterior": rec.get("leitura_anterior"),
        "leitura_atual":    rec.get("leitura_atual"),
        "preco_tusd":     rec.get("preco_tusd"),
        "valor_tusd":     rec.get("valor_tusd"),
        "preco_te":       rec.get("preco_te"),
        "valor_te":       rec.get("valor_te"),
        "tarifa_tusd_sem": rec.get("tarifa_tusd_sem_trib"),
        "tarifa_te_sem":   rec.get("tarifa_te_sem_trib"),
        "bandeira":       rec.get("bandeira_cor"),
        "valor_bandeira": rec.get("valor_bandeira"),
        "cosip":          rec.get("cosip"),
        "total_fatura":   rec.get("total_fatura"),
        "icms_base":      rec.get("icms_base"),
        "icms_aliq":      rec.get("icms_aliq"),
        "icms_valor":     rec.get("icms_valor"),
        "pis_aliq":       rec.get("pis_aliq"),
        "pis_valor":      rec.get("pis_valor"),
        "cofins_aliq":    rec.get("cofins_aliq"),
        "cofins_valor":   rec.get("cofins_valor"),
        "__triagem__":       triagem,
        "__motivos__":       " | ".join(motivos) if motivos else "",
        "__dif_tusd__":      metricas.get("dif_TUSD_R$"),
        "__dif_te__":        metricas.get("dif_TE_R$"),
        "__dif_band__":      None,
        "__dif_leit__":      metricas.get("dif_leit_kWh"),
        "__dif_icms__":      metricas.get("dif_ICMS_R$"),
        "__dif_total__":     metricas.get("dif_total_R$"),
        "__dif_total_pct__": metricas.get("dif_total_%"),
    }


def _triagem_cpfl(alertas):
    cats_criticas = {"Tarifa TUSD", "Tarifa TE", "Bandeira"}
    flags_div = []
    flags_inv = []

    for a in alertas:
        st  = a.get("status", "OK")
        cat = a.get("cat", "")
        if st == "INVESTIGAR":
            if cat in cats_criticas:
                dif = abs(a.get("diferenca") or 0)
                if dif >= 1.0:
                    flags_div.append("[" + cat + "] " + a.get("descricao", ""))
                else:
                    flags_inv.append("[" + cat + "] " + a.get("descricao", ""))
            else:
                flags_inv.append("[" + cat + "] " + a.get("descricao", ""))
        elif st in ("ATENCAO", "ATENCAO"):
            flags_inv.append("[" + cat + "] " + a.get("descricao", ""))

    if flags_div:
        return "DIVERGENCIA", flags_div + flags_inv
    if flags_inv:
        return "INVESTIGAR", flags_inv
    return "OK", []


def normalizar_cpfl(dados, audit_result, pdf_filename):
    alertas  = audit_result.get("alertas", [])
    auditado = audit_result.get("auditado", {})
    itens    = dados.get("itens", [])
    trib     = dados.get("tributos", {})

    triagem, motivos = _triagem_cpfl(alertas)

    tusd_item = next((i for i in itens if i.get("tipo") == "consumo_tusd"), {})
    te_item   = next((i for i in itens if i.get("tipo") == "consumo_te"), {})
    band_val  = sum(i.get("valor") or 0 for i in itens if i.get("tipo") == "bandeira")

    medidor = (dados.get("medidores") or [{}])[0]

    def _dif_cat(cat):
        a = next((x for x in alertas if x.get("cat") == cat), None)
        return a.get("diferenca") if a else None

    consumo = dados.get("consumo_faturado") or 0
    dif_tusd_tar = _dif_cat("Tarifa TUSD")
    dif_te_tar   = _dif_cat("Tarifa TE")
    dif_tusd_rs  = round(dif_tusd_tar * consumo, 2) if dif_tusd_tar and consumo else None
    dif_te_rs    = round(dif_te_tar * consumo, 2) if dif_te_tar and consumo else None
    dif_band_rs  = _dif_cat("Bandeira")

    dif_leit = None
    for a in alertas:
        if a.get("cat") == "Consumo Medidor":
            dif_leit = a.get("diferenca")

    dif_total_rs  = _dif_cat("Total a Pagar")
    total_fat     = dados.get("total_fatura") or 0
    dif_total_pct = round(abs(dif_total_rs) / total_fat * 100, 1) if dif_total_rs and total_fat else None

    return {
        "arquivo":        pdf_filename,
        "distribuidora":  "CPFL Piratininga",
        "layout":         dados.get("_formato"),
        "ref_mes_ano":    dados.get("mes_ref"),
        "vencimento":     _fmt_date(dados.get("data_vencimento")),
        "conta_uc":       dados.get("conta_contrato") or dados.get("uc"),
        "cliente_nome":   dados.get("cliente_nome"),
        "subgrupo":       dados.get("subgrupo"),
        "data_emissao":   _fmt_date(dados.get("data_emissao")),
        "nr_nota_fiscal": dados.get("nota_fiscal"),
        "nr_medidor":     medidor.get("numero") or dados.get("uc"),
        "consumo_kwh":    consumo,
        "nr_dias":        dados.get("dias_ciclo"),
        "leitura_anterior": medidor.get("leitura_anterior"),
        "leitura_atual":    medidor.get("leitura_atual"),
        "preco_tusd":     tusd_item.get("preco_unit_com_trib"),
        "valor_tusd":     tusd_item.get("valor"),
        "preco_te":       te_item.get("preco_unit_com_trib"),
        "valor_te":       te_item.get("valor"),
        "tarifa_tusd_sem": auditado.get("tusd_sem_trib"),
        "tarifa_te_sem":   auditado.get("te_sem_trib"),
        "bandeira":       dados.get("bandeira_vigente"),
        "valor_bandeira": band_val or None,
        "cosip":          None,
        "total_fatura":   total_fat,
        "icms_base":      trib.get("icms", {}).get("base"),
        "icms_aliq":      trib.get("icms", {}).get("aliquota_pct"),
        "icms_valor":     trib.get("icms", {}).get("valor"),
        "pis_aliq":       trib.get("pis", {}).get("aliquota_pct"),
        "pis_valor":      trib.get("pis", {}).get("valor"),
        "cofins_aliq":    trib.get("cofins", {}).get("aliquota_pct"),
        "cofins_valor":   trib.get("cofins", {}).get("valor"),
        "__triagem__":       triagem,
        "__motivos__":       " | ".join(motivos) if motivos else "",
        "__dif_tusd__":      dif_tusd_rs,
        "__dif_te__":        dif_te_rs,
        "__dif_band__":      dif_band_rs,
        "__dif_leit__":      dif_leit,
        "__dif_icms__":      None,
        "__dif_total__":     dif_total_rs,
        "__dif_total_pct__": dif_total_pct,
    }


# -- Processadores ------------------------------------------------------------

def processar_neoenergia_pe(pasta_pdfs):
    from neoenergia_pe.extractor import parse_fatura
    from neoenergia_pe.audit import auditar

    pdfs = sorted(Path(pasta_pdfs).glob("*.pdf"))
    print("  Neoenergia PE: " + str(len(pdfs)) + " PDFs em " + str(pasta_pdfs))
    registros = []
    for pdf in pdfs:
        print("    " + pdf.name + " ...", end=" ", flush=True)
        r = parse_fatura(str(pdf))
        triagem, motivos, metricas = auditar(r)
        rec = normalizar_neoenergia_pe(r, triagem, motivos, metricas)
        registros.append(rec)
        print(rec["__triagem__"])
    return registros


def processar_cpfl(pasta_pdfs, config=None):
    sys.path.insert(0, str(ROOT / "app"))
    from extractor import extract_fatura
    from audit import auditar_fatura

    if config is None:
        config = dict(CPFL_CONFIG_DEFAULT)

    pdfs = sorted(Path(pasta_pdfs).glob("*.pdf"))
    print("  CPFL Piratininga: " + str(len(pdfs)) + " PDFs em " + str(pasta_pdfs))
    registros = []
    seen_nf = set()
    for pdf in pdfs:
        print("    " + pdf.name + " ...", end=" ", flush=True)
        dados = extract_fatura(str(pdf))
        nf = dados.get("nota_fiscal", "")
        if nf and nf in seen_nf:
            print("DUP (ignorado)")
            continue
        if nf:
            seen_nf.add(nf)
        cfg = dict(config)
        cfg["tem_gd"] = dados.get("tem_gd", False)
        if cfg.get("usar_cat_como_compensada", True):
            cfg["energia_compensada_kwh"] = (
                dados.get("gd_ajuste_cat") or dados.get("gd_injetada_mes")
            )
        audit_result = auditar_fatura(dados, cfg)
        rec = normalizar_cpfl(dados, audit_result, pdf.name)
        registros.append(rec)
        print(rec["__triagem__"])
    return registros


# -- Excel individual ---------------------------------------------------------

def _gerar_individual(rec, pasta_out):
    from openpyxl import Workbook as WB
    from core.excel_mestre import (
        COR, _fill, _font, _border, _align, FMT_BRL, COLS,
        _bg_triagem, _fg_triagem
    )

    os.makedirs(pasta_out, exist_ok=True)
    arq = rec["arquivo"].replace(".pdf", "")
    tri = rec["__triagem__"]

    wb = WB()
    ws = wb.active
    ws.title = "Fatura"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24

    ws["A1"] = "FATURA " + arq.upper() + " | " + str(rec.get("ref_mes_ano", ""))
    ws["A1"].font = _font(bold=True, color=COR["header_fg"], size=12)
    ws["A1"].fill = _fill(COR["header_bg"])
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = _align("center")

    ws["A2"] = "TRIAGEM"
    ws["B2"] = tri
    ws["A2"].font = _font(bold=True, size=10)
    ws["B2"].font = _font(bold=True, color=_fg_triagem(tri), size=11)
    ws["B2"].fill = _fill(_bg_triagem(tri))
    ws["A2"].fill = _fill(_bg_triagem(tri))

    mot = rec.get("__motivos__", "")
    if mot:
        ws["A3"] = "Motivos"
        ws["B3"] = mot
        ws["B3"].alignment = _align(wrap=True)
        ws.row_dimensions[3].height = max(30, 14 * len(mot.split("|")))

    r = 5
    for cab, key, fmt, _ in COLS:
        if key.startswith("__"):
            continue
        val = rec.get(key)
        ws.cell(r, 1, cab).font = _font(bold=True, size=9, color="444444")
        c = ws.cell(r, 2, val)
        c.font = _font(size=9)
        if fmt != "@" and val is not None:
            c.number_format = fmt
        r += 1

    fname = os.path.join(pasta_out, "fatura_" + arq + ".xlsx")
    wb.save(fname)


# -- COSIP check --------------------------------------------------------------

def aplicar_check_cosip(registros, tolerancia=0.10):
    from collections import defaultdict
    uc_cosips = defaultdict(list)
    for reg in registros:
        uc = reg.get("conta_uc")
        cosip = reg.get("cosip")
        if uc and cosip and cosip > 0:
            uc_cosips[uc].append(cosip)
    for reg in registros:
        uc = reg.get("conta_uc")
        cosip = reg.get("cosip")
        if not uc or not cosip or cosip <= 0:
            continue
        valores = uc_cosips.get(uc, [])
        if len(valores) < 2:
            continue
        media = sum(valores) / len(valores)
        limite = round(media * (1 + tolerancia), 2)
        if cosip > limite:
            if reg.get("__triagem__") != "DIVERGENCIA":
                reg["__triagem__"] = "INVESTIGAR"
            motivo = f"COSIP R${cosip:.2f} > +10% da media historica UC (media R${media:.2f}, limite R${limite:.2f})"
            atual = reg.get("__motivos__", "")
            reg["__motivos__"] = (atual + " | " + motivo).lstrip(" | ") if atual else motivo
    return registros


# -- Main ---------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Auditoria unificada de faturas de energia")
    parser.add_argument("--concessionaria", choices=["neoenergia_pe", "cpfl_piratininga", "todas"],
                        default="todas")
    parser.add_argument("--pasta",      default=None)
    parser.add_argument("--pasta_neo",  default=None)
    parser.add_argument("--pasta_cpfl", default=None)
    parser.add_argument("--data_adesao_mmgd", default="2022-01-01")
    parser.add_argument("--sem_gd", action="store_true")
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    run_id  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.output) if args.output else ROOT / "output" / ("run_" + run_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    ind_dir = out_dir / "individuais"
    ind_dir.mkdir(exist_ok=True)

    cpfl_config = dict(CPFL_CONFIG_DEFAULT)
    cpfl_config["data_adesao_mmgd"] = args.data_adesao_mmgd

    registros = []

    print("\n" + "=" * 60)
    print("  AUDITORIA UNIFICADA -- " + run_id)
    print("=" * 60)

    if args.concessionaria in ("neoenergia_pe", "todas"):
        pasta = args.pasta_neo or args.pasta
        if not pasta:
            print("  [AVISO] Neoenergia PE: informe --pasta ou --pasta_neo")
        else:
            registros += processar_neoenergia_pe(pasta)

    if args.concessionaria in ("cpfl_piratininga", "todas"):
        pasta = args.pasta_cpfl or args.pasta
        if not pasta:
            print("  [AVISO] CPFL: informe --pasta ou --pasta_cpfl")
        else:
            registros += processar_cpfl(pasta, cpfl_config)

    if not registros:
        print("Nenhuma fatura processada. Verifique os argumentos.")
        sys.exit(1)

    print("\n  Gerando " + str(len(registros)) + " Excel individuais...")
    for rec in registros:
        _gerar_individual(rec, str(ind_dir))

    print("  Gerando Excel-mestre...")
    from core.excel_mestre import gerar_excel_mestre
    master_path = out_dir / ("auditoria_mestre_" + run_id + ".xlsx")
    registros = aplicar_check_cosip(registros)
    gerar_excel_mestre(registros, str(master_path))

    oks   = sum(1 for r in registros if r["__triagem__"] == "OK")
    invs  = sum(1 for r in registros if r["__triagem__"] == "INVESTIGAR")
    divs  = sum(1 for r in registros if r["__triagem__"] == "DIVERGENCIA")
    total = sum(r.get("total_fatura") or 0 for r in registros)

    print("\n" + "=" * 60)
    print("  RESULTADO: " + str(len(registros)) + " faturas | OK=" + str(oks) + " | INVESTIGAR=" + str(invs) + " | DIVERGENCIA=" + str(divs))
    print("  Valor total: R$ " + "{:,.2f}".format(total))
    print("  Saida: " + str(master_path))
    print("=" * 60 + "\n")
    return str(master_path)


if __name__ == "__main__":
    main()
