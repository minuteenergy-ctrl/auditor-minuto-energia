"""
Gera o Excel-mestre com tabela resumo de todas as faturas auditadas em batch.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Arial", size=11, bold=True)
FONT_NORMAL = Font(name="Arial", size=10)
FONT_LINK = Font(name="Arial", size=10, color="0563C1", underline="single")

FILL_TITLE = PatternFill("solid", start_color="1F4E78")
FILL_HEADER = PatternFill("solid", start_color="DDEBF7")
FILL_OK = PatternFill("solid", start_color="E2EFDA")
FILL_ATENCAO = PatternFill("solid", start_color="FCE4D6")
FILL_INVESTIGAR = PatternFill("solid", start_color="FFF2CC")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_BR_RS = '"R$ "#,##0.00;[Red]("R$ "#,##0.00);-'


def gerar_master(resultados, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value="AUDITORIA EM LOTE — Faturas CPFL Piratininga").font = FONT_TITLE
    ws.cell(row=1, column=1).fill = FILL_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ws.row_dimensions[1].height = 28

    headers = [
        "#", "Arquivo PDF", "Cliente", "UC", "Subgrupo",
        "Periodo", "Dias", "Total Fatura (R$)", "Total a Pagar (R$)",
        "OK", "ATENCAO", "INVESTIGAR", "REH aplicada", "Auditoria detalhada",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER

    widths = [4, 32, 28, 12, 10, 24, 6, 15, 15, 7, 11, 13, 13, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 4
    for idx, item in enumerate(resultados, 1):
        dados = item.get("dados", {})
        audit = item.get("audit_result", {})
        alertas = audit.get("alertas", [])
        counts = {"OK": 0, "ATENÇÃO": 0, "INVESTIGAR": 0}
        for a in alertas:
            s = a.get("status", "OK")
            counts[s] = counts.get(s, 0) + 1

        ws.cell(row=r, column=1, value=idx).alignment = ALIGN_CENTER
        ws.cell(row=r, column=2, value=item.get("pdf_filename", "")).font = FONT_NORMAL
        ws.cell(row=r, column=3, value=dados.get("cliente_nome", "")).font = FONT_NORMAL
        ws.cell(row=r, column=4, value=dados.get("uc", "")).alignment = ALIGN_CENTER
        ws.cell(row=r, column=5, value=dados.get("subgrupo", "")).alignment = ALIGN_CENTER

        if dados.get("leitura_anterior") and dados.get("leitura_atual"):
            la = dados["leitura_anterior"].strftime("%d/%m/%Y")
            lt = dados["leitura_atual"].strftime("%d/%m/%Y")
            ws.cell(row=r, column=6, value=f"{la} a {lt}").alignment = ALIGN_CENTER
        ws.cell(row=r, column=7, value=dados.get("dias_ciclo", "")).alignment = ALIGN_CENTER

        c8 = ws.cell(row=r, column=8, value=dados.get("total_fatura", 0))
        c8.number_format = NUM_BR_RS
        c9 = ws.cell(row=r, column=9, value=dados.get("total_a_pagar", 0))
        c9.number_format = NUM_BR_RS

        c10 = ws.cell(row=r, column=10, value=counts["OK"])
        if counts["OK"] > 0:
            c10.fill = FILL_OK
        c10.alignment = ALIGN_CENTER
        c11 = ws.cell(row=r, column=11, value=counts["ATENÇÃO"])
        if counts["ATENÇÃO"] > 0:
            c11.fill = FILL_ATENCAO
        c11.alignment = ALIGN_CENTER
        c12 = ws.cell(row=r, column=12, value=counts["INVESTIGAR"])
        if counts["INVESTIGAR"] > 0:
            c12.fill = FILL_INVESTIGAR
        c12.alignment = ALIGN_CENTER

        ws.cell(row=r, column=13, value=audit.get("auditado", {}).get("reh_aplicada", "")).alignment = ALIGN_CENTER

        link = item.get("excel_individual_path", "")
        if link:
            cell = ws.cell(row=r, column=14, value=Path(link).name)
            cell.hyperlink = link
            cell.font = FONT_LINK
            cell.alignment = ALIGN_LEFT

        r += 1

    r += 1
    ws.cell(row=r, column=1, value="TOTAIS").font = FONT_HEADER
    ws.cell(row=r, column=8, value=f"=SUM(H4:H{r-2})").number_format = NUM_BR_RS
    ws.cell(row=r, column=8).font = FONT_HEADER
    ws.cell(row=r, column=10, value=f"=SUM(J4:J{r-2})").alignment = ALIGN_CENTER
    ws.cell(row=r, column=11, value=f"=SUM(K4:K{r-2})").alignment = ALIGN_CENTER
    ws.cell(row=r, column=12, value=f"=SUM(L4:L{r-2})").alignment = ALIGN_CENTER

    ws.auto_filter.ref = f"A3:N{r-2}"
    ws.freeze_panes = "A4"

    # Aba 2: Alertas detalhados
    ws2 = wb.create_sheet("Alertas Detalhados")
    ws2.sheet_view.showGridLines = False

    ws2.cell(row=1, column=1, value="ALERTAS POR FATURA").font = FONT_TITLE
    ws2.cell(row=1, column=1).fill = FILL_TITLE
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    h2 = ["#", "Arquivo", "Categoria", "Descricao", "Status", "Diferenca/Detalhe"]
    for i, h in enumerate(h2, 1):
        c = ws2.cell(row=3, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER

    w2 = [4, 32, 22, 50, 14, 25]
    for i, w in enumerate(w2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    r = 4
    for idx, item in enumerate(resultados, 1):
        for a in item.get("audit_result", {}).get("alertas", []):
            ws2.cell(row=r, column=1, value=idx).alignment = ALIGN_CENTER
            ws2.cell(row=r, column=2, value=item.get("pdf_filename", ""))
            ws2.cell(row=r, column=3, value=a.get("cat", ""))
            ws2.cell(row=r, column=4, value=a.get("descricao", ""))
            cs = ws2.cell(row=r, column=5, value=a.get("status", ""))
            cs.alignment = ALIGN_CENTER
            cs.font = FONT_HEADER
            if a.get("status") == "OK":
                cs.fill = FILL_OK
            elif a.get("status") == "ATENÇÃO":
                cs.fill = FILL_ATENCAO
            elif a.get("status") == "INVESTIGAR":
                cs.fill = FILL_INVESTIGAR
            extras = []
            if "diferenca" in a:
                extras.append(f"diff={a['diferenca']}")
            if "valor" in a:
                extras.append(f"valor={a['valor']}")
            if "valor_total" in a:
                extras.append(f"total={a['valor_total']}")
            if "compensada" in a:
                extras.append(f"comp={a['compensada']}, inj={a.get('injetada', 0)}")
            ws2.cell(row=r, column=6, value=" | ".join(extras)).alignment = ALIGN_LEFT
            r += 1

    ws2.auto_filter.ref = f"A3:F{r-1}"
    ws2.freeze_panes = "A4"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
