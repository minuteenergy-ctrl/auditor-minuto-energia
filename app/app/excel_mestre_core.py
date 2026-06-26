# -*- coding: utf-8 -*-
"""
core/excel_mestre.py
Geracao de Excel-mestre unificado para todas as concessionarias.
Formato: 5 abas — RESUMO / MESTRE / OK / INVESTIGAR / DIVERGENCIA
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────
COR = {
    "header_bg":   "1F4E79",
    "header_fg":   "FFFFFF",
    "ok_bg":       "E2EFDA",
    "inv_bg":      "FFF2CC",
    "div_bg":      "FCE4D6",
    "alt_bg":      "F2F2F2",
    "branco":      "FFFFFF",
    "triagem_ok":  "375623",
    "triagem_inv": "7F6000",
    "triagem_div": "843C0C",
    "tab_ok":      "70AD47",
    "tab_inv":     "FFD966",
    "tab_div":     "FF7043",
}

def _fill(hex_):
    return PatternFill("solid", start_color=hex_, end_color=hex_)

def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

FMT_BRL  = "#,##0.00"
FMT_TAR  = "0.00000000"
FMT_KWH  = "#,##0.000"
FMT_PCT  = "0.00"
FMT_TEXT = "@"

# ── Schema de colunas comuns ──────────────────────────────────────────────────
# (cabecalho, chave_no_registro, formato, largura)
COLS = [
    ("Arquivo",           "arquivo",           FMT_TEXT, 12),
    ("Distribuidora",     "distribuidora",     FMT_TEXT, 14),
    ("Layout/Formato",    "layout",            FMT_TEXT,  9),
    ("Ref Mes/Ano",       "ref_mes_ano",       FMT_TEXT,  9),
    ("Vencimento",        "vencimento",        FMT_TEXT, 11),
    ("Conta/UC",          "conta_uc",          FMT_TEXT, 14),
    ("Cliente",           "cliente_nome",      FMT_TEXT, 22),
    ("Subgrupo",          "subgrupo",          FMT_TEXT,  8),
    ("Data Emissao",      "data_emissao",      FMT_TEXT, 11),
    ("Nr Nota Fiscal",    "nr_nota_fiscal",    FMT_TEXT, 17),
    ("Nr Medidor",        "nr_medidor",        FMT_TEXT, 12),
    ("Consumo kWh",       "consumo_kwh",       "#,##0.0", 10),
    ("Nr Dias",           "nr_dias",           "#,##0",    7),
    ("Leit Anterior",     "leitura_anterior",  "#,##0.0", 12),
    ("Leit Atual",        "leitura_atual",     "#,##0.0", 10),
    ("Preco TUSD c/trib", "preco_tusd",        FMT_TAR,  15),
    ("Valor TUSD (R$)",   "valor_tusd",        FMT_BRL,  12),
    ("Preco TE c/trib",   "preco_te",          FMT_TAR,  14),
    ("Valor TE (R$)",     "valor_te",          FMT_BRL,  11),
    ("TUSD sem trib",     "tarifa_tusd_sem",   FMT_TAR,  13),
    ("TE sem trib",       "tarifa_te_sem",     FMT_TAR,  12),
    ("Bandeira",          "bandeira",          FMT_TEXT, 12),
    ("Val Bandeira (R$)", "valor_bandeira",    FMT_BRL,  13),
    ("COSIP (R$)",        "cosip",             FMT_BRL,  10),
    ("Total Fatura (R$)", "total_fatura",      FMT_BRL,  14),
    ("ICMS Base (R$)",    "icms_base",         FMT_BRL,  12),
    ("ICMS Aliq %",       "icms_aliq",         FMT_PCT,  10),
    ("ICMS Valor (R$)",   "icms_valor",        FMT_BRL,  12),
    ("PIS Aliq %",        "pis_aliq",          FMT_PCT,  10),
    ("PIS Valor (R$)",    "pis_valor",         FMT_BRL,  10),
    ("COFINS Aliq %",     "cofins_aliq",       FMT_PCT,  12),
    ("COFINS Valor (R$)", "cofins_valor",      FMT_BRL,  13),
    # ── Auditoria ──
    ("TRIAGEM",           "__triagem__",       FMT_TEXT, 11),
    ("Motivos / Alertas", "__motivos__",       FMT_TEXT, 50),
    ("Dif TUSD (R$)",     "__dif_tusd__",      FMT_BRL,  12),
    ("Dif TE (R$)",       "__dif_te__",        FMT_BRL,  11),
    ("Dif Bandeira (R$)", "__dif_band__",      FMT_BRL,  13),
    ("Dif Leit kWh",      "__dif_leit__",      "#,##0.0", 11),
    ("Dif ICMS (R$)",     "__dif_icms__",      FMT_BRL,  11),
    ("Dif Total (R$)",    "__dif_total__",     FMT_BRL,  12),
    ("Dif Total %",       "__dif_total_pct__", "0.0",    10),
]


# ── Colunas extras — modelo MT (media tensao com demanda) ────────────────────
COLS_MT_EXTRA = [
    # Faturado MT
    ("Consumo Pta kWh Fat",    "consumo_ponta_kwh",      FMT_KWH,   16),
    ("Consumo FP kWh Fat",     "consumo_fp_kwh",         FMT_KWH,   14),
    ("Demanda Cont kW",        "demanda_contratada_kw",  "#,##0.0", 14),
    ("Demanda Med kW",         "demanda_medida_kw",      "#,##0.0", 13),
    ("Demanda Ultrap kW",      "demanda_ultrap_kw",      "#,##0.0", 13),
    ("Taxa Perda",             "taxa_perda",             FMT_TEXT,   9),
    # kWh Ponta
    ("kWh Pta L.Ant",          "med_kwh_ponta_lant",     "#,##0",   11),
    ("kWh Pta L.Atu",          "med_kwh_ponta_latu",     "#,##0",   11),
    ("kWh Pta Mult",           "med_kwh_ponta_mult",     "0.00000",  9),
    ("kWh Pta Medido",         "med_kwh_ponta_cons",     FMT_KWH,   13),
    # kWh Fora Ponta
    ("kWh FP L.Ant",           "med_kwh_fp_lant",        "#,##0",   11),
    ("kWh FP L.Atu",           "med_kwh_fp_latu",        "#,##0",   11),
    ("kWh FP Mult",            "med_kwh_fp_mult",        "0.00000",  9),
    ("kWh FP Medido",          "med_kwh_fp_cons",        FMT_KWH,   13),
    # kW Ponta
    ("kW Pta L.Ant",           "med_kw_ponta_lant",      "#,##0",   10),
    ("kW Pta L.Atu",           "med_kw_ponta_latu",      "#,##0",   10),
    ("kW Pta Mult",            "med_kw_ponta_mult",      "0.00000",  9),
    ("kW Pta Medido",          "med_kw_ponta_cons",      "#,##0.0", 12),
    # kW Fora Ponta
    ("kW FP L.Ant",            "med_kw_fp_lant",         "#,##0",   10),
    ("kW FP L.Atu",            "med_kw_fp_latu",         "#,##0",   10),
    ("kW FP Mult",             "med_kw_fp_mult",         "0.00000",  9),
    ("kW FP Medido",           "med_kw_fp_cons",         "#,##0.0", 12),
    # kVarh Ponta
    ("kVarh Pta L.Ant",        "med_kvarh_ponta_lant",   "#,##0",   12),
    ("kVarh Pta L.Atu",        "med_kvarh_ponta_latu",   "#,##0",   12),
    ("kVarh Pta Mult",         "med_kvarh_ponta_mult",   "0.00000", 10),
    ("kVarh Pta Medido",       "med_kvarh_ponta_cons",   FMT_KWH,   13),
    # kVarh Fora Ponta
    ("kVarh FP L.Ant",         "med_kvarh_fp_lant",      "#,##0",   11),
    ("kVarh FP L.Atu",         "med_kvarh_fp_latu",      "#,##0",   11),
    ("kVarh FP Mult",          "med_kvarh_fp_mult",      "0.00000", 10),
    ("kVarh FP Medido",        "med_kvarh_fp_cons",      FMT_KWH,   13),
]

COLS_MT = COLS + COLS_MT_EXTRA


def _bg_triagem(t):
    return COR["div_bg"] if t == "DIVERGENCIA" else COR["inv_bg"] if t == "INVESTIGAR" else COR["ok_bg"]

def _fg_triagem(t):
    return COR["triagem_div"] if t == "DIVERGENCIA" else COR["triagem_inv"] if t == "INVESTIGAR" else COR["triagem_ok"]


def _escrever_cabecalho(ws, cols):
    for ci, (cab, *_) in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=cab)
        c.font      = _font(bold=True, color=COR["header_fg"], size=9)
        c.fill      = _fill(COR["header_bg"])
        c.alignment = _align("center")
        c.border    = _border()
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _escrever_linha(ws, row_num, rec, alt=False, cols=None):
    if cols is None:
        cols = COLS
    triagem = rec.get("__triagem__", "OK")
    bg_row  = COR["alt_bg"] if alt else COR["branco"]
    bg_tri  = _bg_triagem(triagem)
    fg_tri  = _fg_triagem(triagem)

    for ci, (_, key, fmt, _) in enumerate(cols, 1):
        val  = rec.get(key)
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.border    = _border()

        if key == "__triagem__":
            cell.fill      = _fill(bg_tri)
            cell.font      = _font(bold=True, color=fg_tri, size=9)
            cell.alignment = _align("center")
        elif key == "__motivos__":
            cell.fill      = _fill(bg_tri)
            cell.font      = _font(color="444444", size=8)
            cell.alignment = _align("left", wrap=True)
        else:
            cell.fill      = _fill(bg_row)
            cell.font      = _font(size=9)
            cell.alignment = _align("left")

        if fmt != FMT_TEXT and val is not None:
            cell.number_format = fmt


def _ajustar_colunas(ws, cols):
    for ci, (_, _, _, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri in range(2, ws.max_row + 1):
        ws.row_dimensions[ri].height = 16


def _aba_dados(wb, nome, registros, cor_tab=None, cols=None):
    if cols is None:
        cols = COLS
    ws = wb.create_sheet(nome)
    if cor_tab:
        ws.sheet_properties.tabColor = cor_tab
    _escrever_cabecalho(ws, cols)
    for i, rec in enumerate(registros):
        _escrever_linha(ws, i + 2, rec, alt=(i % 2 == 1), cols=cols)
    _ajustar_colunas(ws, cols)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    return ws


def _aba_resumo(wb, meta, contagens, totais):
    ws = wb.create_sheet("RESUMO")
    ws.sheet_properties.tabColor = COR["header_bg"]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    # Titulo
    ws["A1"] = "AUDITORIA DE FATURAS — MINUTO ENERGIA"
    ws["A1"].font = _font(bold=True, color=COR["header_fg"], size=13)
    ws["A1"].fill = _fill(COR["header_bg"])
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 28

    def row(r, a, b=None, c=None, bold=False, bg=None):
        for ci, v in enumerate([a, b, c], 1):
            if v is None:
                continue
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font      = _font(bold=bold, size=10)
            cell.alignment = _align("right" if ci > 1 else "left")
            if bg:
                cell.fill = _fill(bg)

    row(3, "Concessionaria(s):", meta.get("distribuidoras", ""))
    row(4, "Periodo:",           meta.get("periodo", ""))
    row(5, "Total de Faturas:",  meta.get("total", 0))

    row(7, "TRIAGEM", "Qtd", "R$ Total", bold=True, bg=COR["header_bg"])
    ws.cell(7,1).font = _font(bold=True, color=COR["header_fg"])
    ws.cell(7,2).font = _font(bold=True, color=COR["header_fg"])
    ws.cell(7,3).font = _font(bold=True, color=COR["header_fg"])

    for r_idx, (tag, bg, fg) in enumerate([
        ("OK",          COR["ok_bg"],  COR["triagem_ok"]),
        ("INVESTIGAR",  COR["inv_bg"], COR["triagem_inv"]),
        ("DIVERGENCIA", COR["div_bg"], COR["triagem_div"]),
    ], 8):
        qtd = contagens.get(tag, 0)
        tot = totais.get(tag, 0.0)
        ws.cell(r_idx, 1, tag).fill  = _fill(bg)
        ws.cell(r_idx, 1).font       = _font(bold=True, color=fg)
        ws.cell(r_idx, 2, qtd).fill  = _fill(bg)
        ws.cell(r_idx, 2).alignment  = _align("right")
        ws.cell(r_idx, 3, tot).fill  = _fill(bg)
        ws.cell(r_idx, 3).number_format = FMT_BRL
        ws.cell(r_idx, 3).alignment  = _align("right")

    r_tot = 11
    ws.cell(r_tot, 1, "TOTAL").font = _font(bold=True)
    ws.cell(r_tot, 2, contagens.get("TOTAL", 0)).font = _font(bold=True)
    ws.cell(r_tot, 2).alignment = _align("right")
    ws.cell(r_tot, 3, totais.get("TOTAL", 0.0)).font = _font(bold=True)
    ws.cell(r_tot, 3).number_format = FMT_BRL
    ws.cell(r_tot, 3).alignment = _align("right")
    return ws


def gerar_excel_mestre(registros, output_path, modelo=None):
    """
    Gera o Excel-mestre unificado.
    registros : lista de dicts com schema padrao (ver COLS acima).
    output_path: caminho do arquivo .xlsx a salvar.
    modelo    : "BT" | "MT" | None (auto-detecta pelo subgrupo).
                Subgrupos A*, A3, A4, A3a etc. → MT (colunas de leitura por posto).
    """
    if modelo is None:
        subgrupos = {r.get("subgrupo", "") for r in registros if r.get("subgrupo")}
        modelo = "MT" if any(str(s).upper().startswith("A") for s in subgrupos) else "BT"
    cols = COLS_MT if modelo == "MT" else COLS

    oks   = [r for r in registros if r.get("__triagem__") == "OK"]
    invs  = [r for r in registros if r.get("__triagem__") == "INVESTIGAR"]
    divs  = [r for r in registros if r.get("__triagem__") == "DIVERGENCIA"]

    contagens = {
        "OK": len(oks), "INVESTIGAR": len(invs),
        "DIVERGENCIA": len(divs), "TOTAL": len(registros),
    }
    totais = {
        "OK":          round(sum(r.get("total_fatura") or 0 for r in oks), 2),
        "INVESTIGAR":  round(sum(r.get("total_fatura") or 0 for r in invs), 2),
        "DIVERGENCIA": round(sum(r.get("total_fatura") or 0 for r in divs), 2),
        "TOTAL":       round(sum(r.get("total_fatura") or 0 for r in registros), 2),
    }

    dists = sorted({r.get("distribuidora", "") for r in registros if r.get("distribuidora")})
    refs  = sorted(r.get("ref_mes_ano", "") for r in registros if r.get("ref_mes_ano"))
    meta  = {
        "distribuidoras": " | ".join(dists),
        "periodo": f"{refs[0]} a {refs[-1]}" if refs else "",
        "total": len(registros),
    }

    wb = Workbook()
    wb.remove(wb.active)

    _aba_resumo(wb, meta, contagens, totais)
    _aba_dados(wb, "MESTRE",      registros, cor_tab=COR["header_bg"], cols=cols)
    _aba_dados(wb, "OK",          oks,       cor_tab=COR["tab_ok"],    cols=cols)
    _aba_dados(wb, "INVESTIGAR",  invs,      cor_tab=COR["tab_inv"],   cols=cols)
    _aba_dados(wb, "DIVERGENCIA", divs,      cor_tab=COR["tab_div"],   cols=cols)

    from pathlib import Path
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
