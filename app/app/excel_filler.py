"""
Excel Filler — cria o arquivo de auditoria individual para cada fatura.
Nao usa template externo: gera o workbook do zero via openpyxl.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

FONT_TITLE = Font(name="Arial", size=13, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Arial", size=11, bold=True)
FONT_NORMAL = Font(name="Arial", size=10)
FONT_LABEL = Font(name="Arial", size=10, bold=True)
FONT_INPUT = Font(name="Arial", size=10, color="0070C0")

FILL_TITLE = PatternFill("solid", start_color="1F4E78")
FILL_HEADER = PatternFill("solid", start_color="DDEBF7")
FILL_INPUT = PatternFill("solid", start_color="EBF3FB")
FILL_OK = PatternFill("solid", start_color="E2EFDA")
FILL_ATENCAO = PatternFill("solid", start_color="FCE4D6")
FILL_INVESTIGAR = PatternFill("solid", start_color="FFF2CC")
FILL_SECTION = PatternFill("solid", start_color="F2F2F2")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_RS = '"R$ "#,##0.00'
NUM_PCT = '0.00%'
NUM_KWH = '#,##0.000'


def _set(ws, row, col, value, font=None, fill=None, align=None, border=None, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border
    if num_format:
        c.number_format = num_format
    return c


def preencher_template(dados, config, audit_result, output_path):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Aba 1: Dados da Fatura ----------------------------------------
    ws = wb.create_sheet("Fatura")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14

    # Titulo
    _set(ws, 1, 2, "AUDITORIA DE FATURA — CPFL Piratininga", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 24

    # Secao: Identificacao
    r = 3
    _set(ws, r, 2, "IDENTIFICAÇÃO", FONT_HEADER, FILL_HEADER, ALIGN_LEFT)
    ws.merge_cells(f"B{r}:E{r}")

    def lbl(row, label, value, num_fmt=None):
        _set(ws, row, 2, label, FONT_LABEL, None, ALIGN_LEFT, BORDER)
        c = _set(ws, row, 3, value, FONT_INPUT, FILL_INPUT, ALIGN_LEFT, BORDER, num_fmt)
        ws.merge_cells(f"C{row}:E{row}")

    r = 4
    lbl(r, "Cliente / Razão Social", dados.get("cliente_nome", "")); r += 1
    lbl(r, "CPF", dados.get("cliente_cpf", "")); r += 1
    lbl(r, "Endereço", dados.get("cliente_endereco", "")); r += 1
    lbl(r, "Unidade Consumidora (UC)", dados.get("uc", "")); r += 1
    lbl(r, "Código do Cliente", dados.get("codigo_cliente", "")); r += 1
    lbl(r, "Conta Contrato", dados.get("conta_contrato", "")); r += 1
    lbl(r, "Subgrupo", dados.get("subgrupo", "B1")); r += 1
    lbl(r, "Modalidade Tarifária", dados.get("modalidade", "Convencional")); r += 1
    lbl(r, "Tipo de Fornecimento", dados.get("tipo_fornecimento", "")); r += 1

    da = config.get("data_adesao_mmgd")
    if isinstance(da, str):
        try:
            da = datetime.datetime.strptime(da, "%Y-%m-%d").date()
        except Exception:
            pass
    lbl(r, "Data de Adesão MMGD", da); r += 1
    lbl(r, "Possui GD?", "Sim" if dados.get("tem_gd") else "Não"); r += 1

    r += 1
    _set(ws, r, 2, "FATURA", FONT_HEADER, FILL_HEADER, ALIGN_LEFT)
    ws.merge_cells(f"B{r}:E{r}"); r += 1

    lbl(r, "Mês de Referência", dados.get("mes_ref", "")); r += 1
    lbl(r, "Nota Fiscal", dados.get("nota_fiscal", "")); r += 1
    if dados.get("data_emissao"):
        lbl(r, "Data de Emissão", dados["data_emissao"]); r += 1
    if dados.get("data_vencimento"):
        lbl(r, "Data de Vencimento", dados["data_vencimento"]); r += 1
    if dados.get("leitura_anterior"):
        lbl(r, "Data Leitura Anterior", dados["leitura_anterior"]); r += 1
    if dados.get("leitura_atual"):
        lbl(r, "Data Leitura Atual", dados["leitura_atual"]); r += 1
    lbl(r, "Dias do Ciclo", dados.get("dias_ciclo", "")); r += 1
    lbl(r, "Consumo Faturado (kWh)", dados.get("consumo_faturado", ""), NUM_KWH); r += 1
    lbl(r, "Total da Fatura (R$)", dados.get("total_fatura", 0), NUM_RS); r += 1
    lbl(r, "Total a Pagar (R$)", dados.get("total_a_pagar", 0), NUM_RS); r += 1

    r += 1
    _set(ws, r, 2, "ITENS COBRADOS", FONT_HEADER, FILL_HEADER, ALIGN_LEFT)
    ws.merge_cells(f"B{r}:E{r}"); r += 1

    _set(ws, r, 2, "Item", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER)
    _set(ws, r, 3, "Qtd (kWh)", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER)
    _set(ws, r, 4, "Tarifa (R$/kWh)", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER)
    _set(ws, r, 5, "Valor (R$)", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER); r += 1

    for it in dados.get("itens", []):
        _set(ws, r, 2, it.get("descricao", it.get("tipo", "")), FONT_NORMAL, None, ALIGN_LEFT, BORDER)
        _set(ws, r, 3, it.get("quantidade"), FONT_NORMAL, None, ALIGN_RIGHT, BORDER, NUM_KWH)
        _set(ws, r, 4, it.get("preco_unit_com_trib") or it.get("tarifa_unit_sem"), FONT_NORMAL, None, ALIGN_RIGHT, BORDER, "0.000000")
        _set(ws, r, 5, it.get("valor"), FONT_NORMAL, None, ALIGN_RIGHT, BORDER, NUM_RS)
        r += 1

    r += 1
    _set(ws, r, 2, "TRIBUTOS", FONT_HEADER, FILL_HEADER, ALIGN_LEFT)
    ws.merge_cells(f"B{r}:E{r}"); r += 1

    _set(ws, r, 2, "Tributo", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER)
    _set(ws, r, 3, "Base (R$)", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER)
    _set(ws, r, 4, "Alíquota", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER)
    _set(ws, r, 5, "Valor (R$)", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER); r += 1

    trib = dados.get("tributos", {})
    for nome, key in [("ICMS", "icms"), ("PIS", "pis"), ("COFINS", "cofins")]:
        t = trib.get(key, {})
        _set(ws, r, 2, nome, FONT_NORMAL, None, ALIGN_LEFT, BORDER)
        _set(ws, r, 3, t.get("base"), FONT_NORMAL, None, ALIGN_RIGHT, BORDER, NUM_RS)
        alq = (t.get("aliquota_pct") or 0) / 100
        _set(ws, r, 4, alq, FONT_NORMAL, None, ALIGN_RIGHT, BORDER, NUM_PCT)
        _set(ws, r, 5, t.get("valor"), FONT_NORMAL, None, ALIGN_RIGHT, BORDER, NUM_RS)
        r += 1

    # ---- Aba 2: Auditoria -----------------------------------------------
    ws2 = wb.create_sheet("Auditoria")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 3
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 55
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 20

    _set(ws2, 1, 2, "RESULTADO DA AUDITORIA — CPFL Piratininga", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    ws2.merge_cells("B1:E1")
    ws2.row_dimensions[1].height = 24

    aud = audit_result.get("auditado", {})
    r2 = 3
    _set(ws2, r2, 2, "TARIFAS AUDITADAS", FONT_HEADER, FILL_HEADER, ALIGN_LEFT)
    ws2.merge_cells(f"B{r2}:E{r2}"); r2 += 1

    def lbl2(row, label, value, num_fmt=None):
        _set(ws2, row, 2, label, FONT_LABEL, None, ALIGN_LEFT, BORDER)
        _set(ws2, row, 3, value, FONT_NORMAL, FILL_INPUT, ALIGN_LEFT, BORDER, num_fmt)
        ws2.merge_cells(f"C{row}:E{row}")

    lbl2(r2, "REH Aplicada", aud.get("reh_aplicada", "")); r2 += 1
    lbl2(r2, "TUSD s/ tributos (R$/kWh)", aud.get("tusd_sem_trib"), "0.000000"); r2 += 1
    lbl2(r2, "TE s/ tributos (R$/kWh)", aud.get("te_sem_trib"), "0.000000"); r2 += 1
    lbl2(r2, "TUSD c/ gross-up (R$/kWh)", aud.get("tusd_com_trib"), "0.000000"); r2 += 1
    lbl2(r2, "TE c/ gross-up (R$/kWh)", aud.get("te_com_trib"), "0.000000"); r2 += 1
    lbl2(r2, "Bandeira Auditada (R$)", aud.get("bandeira_auditada"), NUM_RS); r2 += 1
    lbl2(r2, "Fator Fio B", aud.get("fator_fio_b"), NUM_PCT); r2 += 1

    r2 += 1
    _set(ws2, r2, 2, "ALERTAS", FONT_HEADER, FILL_HEADER, ALIGN_LEFT)
    ws2.merge_cells(f"B{r2}:E{r2}"); r2 += 1

    _set(ws2, r2, 2, "Categoria", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER)
    _set(ws2, r2, 3, "Descrição", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER)
    _set(ws2, r2, 4, "Status", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER)
    _set(ws2, r2, 5, "Diferença/Detalhe", FONT_HEADER, FILL_SECTION, ALIGN_CENTER, BORDER)
    r2 += 1

    fill_map = {"OK": FILL_OK, "ATENÇÃO": FILL_ATENCAO, "INVESTIGAR": FILL_INVESTIGAR}
    for a in audit_result.get("alertas", []):
        st = a.get("status", "OK")
        fl = fill_map.get(st, None)
        _set(ws2, r2, 2, a.get("cat", ""), FONT_NORMAL, fl, ALIGN_LEFT, BORDER)
        _set(ws2, r2, 3, a.get("descricao", ""), FONT_NORMAL, fl, ALIGN_LEFT, BORDER)
        cs = _set(ws2, r2, 4, st, FONT_HEADER, fl, ALIGN_CENTER, BORDER)
        extras = []
        if "diferenca" in a:
            extras.append(f"diff=R${a['diferenca']:.4f}")
        if "valor" in a:
            extras.append(f"val={a['valor']}")
        if "valor_total" in a:
            extras.append(f"total=R${a['valor_total']:.2f}")
        _set(ws2, r2, 5, " | ".join(extras), FONT_NORMAL, fl, ALIGN_LEFT, BORDER)
        r2 += 1

    ws2.freeze_panes = "B4"

    wb.save(output_path)
    return output_path
